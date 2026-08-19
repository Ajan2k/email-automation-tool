"""Parse .xlsx / .csv contact files into normalized row dicts."""
import csv
import io

from openpyxl import load_workbook

EXPECTED_COLUMNS = {
    "first_name",
    "last_name",
    "email",
    "company",
    "job_title",
    "website",
    "linkedin",
    "industry",
}

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
ALLOWED_EXTENSIONS = {".xlsx", ".csv"}


def _normalize_header(header: str) -> str:
    return (header or "").strip().lower().replace(" ", "_")


def parse_contact_file(filename: str, content: bytes) -> list[dict]:
    """Return list of row dicts keyed by normalized column names."""
    if len(content) > MAX_FILE_SIZE:
        raise ValueError("File too large (max 10 MB)")

    name = filename.lower()
    if name.endswith(".csv"):
        return _parse_csv(content)
    if name.endswith(".xlsx"):
        return _parse_xlsx(content)
    raise ValueError("Unsupported file type. Upload .xlsx or .csv")


def _parse_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for raw in reader:
        rows.append({_normalize_header(k): (v or "").strip() for k, v in raw.items() if k})
    return rows


def _parse_xlsx(content: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [_normalize_header(str(h) if h is not None else "") for h in next(rows_iter)]
    except StopIteration:
        return []
    rows = []
    for raw in rows_iter:
        if raw is None or all(v is None for v in raw):
            continue
        row = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            value = raw[i] if i < len(raw) else None
            row[header] = str(value).strip() if value is not None else ""
        rows.append(row)
    return rows
